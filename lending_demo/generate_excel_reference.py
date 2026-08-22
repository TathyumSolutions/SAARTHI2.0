#!/usr/bin/env python3
"""
generate_excel_reference.py
============================
Builds the "external data" Excel workbooks that live alongside the lending
demo database - the kind of reference data a real NBFC keeps in
spreadsheets rather than the core loan system, and that Saarthi should be
able to cross-check core DB figures against via its spreadsheet connector
(app/services/spreadsheet_service.py).

Reads product/branch/agent context straight out of the already-loaded
lending_demo_db so the workbooks are consistent with whatever --scale was
used for db_lending.py, then writes 4 workbooks under lending_demo/output/excel/:

  01_rbi_benchmark_rates.xlsx   - RBI policy & benchmark lending rates used
                                  to sanity-check the book's product pricing
  02_branch_master.xlsx         - branch roster + staffing summary
  03_agent_commission_slabs.xlsx- DSA/agent commission slab card by product
  04_insurance_partner_rates.xlsx - credit-linked insurance premium rates

Usage:
    python generate_excel_reference.py --db-url postgresql://saarthi:password@localhost:5432/postgres
"""
from __future__ import annotations

import argparse
import os

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output", "excel")

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="808080")


def derive_db_url(base_url: str, db_name: str) -> str:
    parsed = urlparse(base_url)
    return parsed._replace(path=f"/{db_name}").geturl()


def style_header_row(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, ncols: int, min_width=10, max_width=45):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, max_len + 2)


def write_title_block(ws, title: str, subtitle: str, start_row: int = 1) -> int:
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = NOTE_FONT
    return start_row + 3


# ---------------------------------------------------------------------
# Workbook 1: RBI benchmark / policy rates
# ---------------------------------------------------------------------

def build_rbi_benchmark_workbook(products: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Rates"
    row = write_title_block(
        ws, "RBI Policy & Benchmark Lending Rates - Reference Card",
        "Illustrative reference figures for cross-checking product pricing against policy/market benchmarks. "
        "Not an official RBI publication - synthetic data for demo/testing purposes only.",
    )

    policy_rows = [
        ("Repo Rate", "6.50%", "As last reviewed by RBI's Monetary Policy Committee"),
        ("Standing Deposit Facility (SDF) Rate", "6.25%", ""),
        ("Marginal Standing Facility (MSF) Rate", "6.75%", ""),
        ("Bank Rate", "6.75%", ""),
        ("CRR (Cash Reserve Ratio)", "4.50%", "Of Net Demand & Time Liabilities"),
        ("SLR (Statutory Liquidity Ratio)", "18.00%", ""),
    ]
    ws.cell(row=row, column=1, value="RBI Policy Rates").font = Font(bold=True, size=12)
    row += 1
    headers = ["Rate", "Value", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for r in policy_rows:
        for c, v in enumerate(r, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="NBFC Product Rate Benchmarks vs Book Pricing").font = Font(bold=True, size=12)
    row += 1
    headers = ["Product Code", "Product Name", "Category", "Book Rate Min %", "Book Rate Max %",
               "Market Benchmark Min %", "Market Benchmark Max %", "Within Benchmark?"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    # Market benchmark = book range nudged, to give a realistic "mostly in
    # band, one or two products drifting outside" check for Saarthi demos.
    for p in products:
        bench_min = round(p["rate_min"] - 1.0, 2)
        bench_max = round(p["rate_max"] + 1.0, 2)
        within = "Yes"
        if p["code"] in ("CD", "MFI"):
            bench_max = round(p["rate_max"] - 2.5, 2)  # deliberately flag these two as pricing above market
            within = "Review - above benchmark"
        ws.cell(row=row, column=1, value=p["code"])
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=p["category"])
        ws.cell(row=row, column=4, value=p["rate_min"])
        ws.cell(row=row, column=5, value=p["rate_max"])
        ws.cell(row=row, column=6, value=bench_min)
        ws.cell(row=row, column=7, value=bench_max)
        ws.cell(row=row, column=8, value=within)
        row += 1

    autosize(ws, len(headers))
    return wb


# ---------------------------------------------------------------------
# Workbook 2: Branch master
# ---------------------------------------------------------------------

def build_branch_master_workbook(branches: list[dict], staffing: dict[int, dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Master"
    row = write_title_block(
        ws, "Branch Network Master",
        "Branch roster with staffing summary, maintained outside the core system as the operations team's working file.",
    )
    headers = ["Branch Code", "Branch Name", "City", "State", "Region", "Pincode",
               "Opened Date", "Active?", "Employee Count", "Agent Count"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for b in branches:
        st = staffing.get(b["branch_id"], {"employees": 0, "agents": 0})
        ws.cell(row=row, column=1, value=b["branch_code"])
        ws.cell(row=row, column=2, value=b["branch_name"])
        ws.cell(row=row, column=3, value=b["city"])
        ws.cell(row=row, column=4, value=b["state"])
        ws.cell(row=row, column=5, value=b["region"])
        ws.cell(row=row, column=6, value=b["pincode"])
        ws.cell(row=row, column=7, value=b["opened_date"].isoformat())
        ws.cell(row=row, column=8, value="Yes" if b["is_active"] else "No")
        ws.cell(row=row, column=9, value=st["employees"])
        ws.cell(row=row, column=10, value=st["agents"])
        row += 1
    autosize(ws, len(headers))
    return wb


# ---------------------------------------------------------------------
# Workbook 3: Agent commission slabs
# ---------------------------------------------------------------------

def build_commission_slab_workbook(products: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission Slabs"
    row = write_title_block(
        ws, "DSA / Agent Commission Slab Card",
        "Commission payable to empanelled agents (DSAs) as a % of disbursed amount, by product and by "
        "the agent's trailing-quarter disbursement volume slab.",
    )
    headers = ["Product Code", "Product Name", "Slab: < Rs 10L/qtr", "Slab: Rs 10L-50L/qtr",
               "Slab: Rs 50L-2Cr/qtr", "Slab: > Rs 2Cr/qtr"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for p in products:
        base = (p["rate_min"] + p["rate_max"]) / 2  # not the interest rate - just a seed for a plausible spread
        slab1 = round(0.60 + (base % 3) * 0.05, 2)
        slab2 = round(slab1 + 0.20, 2)
        slab3 = round(slab1 + 0.40, 2)
        slab4 = round(slab1 + 0.55, 2)
        ws.cell(row=row, column=1, value=p["code"])
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=f"{slab1}%")
        ws.cell(row=row, column=4, value=f"{slab2}%")
        ws.cell(row=row, column=5, value=f"{slab3}%")
        ws.cell(row=row, column=6, value=f"{slab4}%")
        row += 1
    row += 2
    ws.cell(row=row, column=1, value="Notes").font = Font(bold=True)
    row += 1
    for note in [
        "Commission is paid on disbursed amount, net of any part-prepayment within the first 90 days.",
        "Slab is re-evaluated at the start of every quarter based on the agent's trailing 3-month disbursement volume.",
        "Gold Loan and Loan Against Property commissions are capped at 1.5% per the product policy regardless of slab.",
    ]:
        ws.cell(row=row, column=1, value=f"- {note}")
        row += 1
    autosize(ws, len(headers))
    return wb


# ---------------------------------------------------------------------
# Workbook 4: Insurance partner rates
# ---------------------------------------------------------------------

def build_insurance_rate_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Insurance Rates"
    row = write_title_block(
        ws, "Credit-Linked Insurance - Partner Rate Card",
        "Premium rates (per Rs 1,000 sum assured per year) charged by bancassurance/insurance partners for "
        "loan-linked group credit life and asset insurance covers.",
    )
    headers = ["Insurance Partner", "Cover Type", "Applicable Loan Categories", "Premium Rate (per Rs 1,000 SA p.a.)",
               "Min Age", "Max Age", "Claim Settlement Ratio %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    rows = [
        ("HDFC Life", "Group Credit Life", "Personal, MSME, LAP, Home", 3.10, 18, 65, 98.7),
        ("ICICI Prudential", "Group Credit Life", "Personal, Vehicle, MSME", 2.95, 18, 60, 97.9),
        ("SBI Life", "Group Credit Life", "Home, LAP, Education", 2.80, 18, 65, 98.5),
        ("Tata AIG", "Asset/Vehicle Insurance", "Vehicle (Two-Wheeler, Used Car, New Car)", 18.50, None, None, 96.8),
        ("Bajaj Allianz", "Asset/Vehicle Insurance", "Vehicle, Consumer Durable", 16.75, None, None, 97.5),
        ("Star Health", "Health Rider (optional add-on)", "Personal, MSME", 4.20, 18, 60, 95.2),
    ]
    for r in rows:
        for c, v in enumerate(r, start=1):
            ws.cell(row=row, column=c, value=v if v is not None else "N/A")
        row += 1
    autosize(ws, len(headers))
    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=os.getenv("LENDING_DATABASE_URL", "postgresql://saarthi:password@localhost:5432/postgres"))
    parser.add_argument("--db-name", default="lending_demo_db")
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = psycopg2.connect(derive_db_url(args.db_url, args.db_name))
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT product_id, product_code AS code, product_name AS name, category,
               interest_rate_min AS rate_min, interest_rate_max AS rate_max
        FROM loan_products ORDER BY product_id
    """)
    products = [dict(r) for r in cur.fetchall()]
    for p in products:
        p["rate_min"] = float(p["rate_min"])
        p["rate_max"] = float(p["rate_max"])

    cur.execute("""
        SELECT branch_id, branch_code, branch_name, city, state, region, pincode, opened_date, is_active
        FROM branches ORDER BY branch_id
    """)
    branches = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT branch_id, count(*) AS n FROM employees GROUP BY branch_id")
    emp_counts = {r["branch_id"]: r["n"] for r in cur.fetchall()}
    cur.execute("SELECT branch_id, count(*) AS n FROM agents GROUP BY branch_id")
    agent_counts = {r["branch_id"]: r["n"] for r in cur.fetchall()}
    staffing = {
        b["branch_id"]: {"employees": emp_counts.get(b["branch_id"], 0), "agents": agent_counts.get(b["branch_id"], 0)}
        for b in branches
    }
    conn.close()

    build_rbi_benchmark_workbook(products).save(os.path.join(OUTPUT_DIR, "01_rbi_benchmark_rates.xlsx"))
    build_branch_master_workbook(branches, staffing).save(os.path.join(OUTPUT_DIR, "02_branch_master.xlsx"))
    build_commission_slab_workbook(products).save(os.path.join(OUTPUT_DIR, "03_agent_commission_slabs.xlsx"))
    build_insurance_rate_workbook().save(os.path.join(OUTPUT_DIR, "04_insurance_partner_rates.xlsx"))

    print(f"Wrote 4 workbooks to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
